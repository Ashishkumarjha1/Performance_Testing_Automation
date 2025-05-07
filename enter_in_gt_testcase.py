# latest_one 7:30 2aug
from openpyxl import load_workbook
import pandas as pd
import mysql.connector
from mysql.connector import Error
import openpyxl
import json
import os


class AlarmSystemDatabase:
    def __init__(self, config_path):
        self.config_path = config_path
        self.connection = None
        # LIst to store the channel names from the testcase table
        self.testcase_channel_and_test_suit_list = []
        # List to store the  TestCaseList from the testrun table
        # self.testcase_channel_list_from_testrun = []
        self.list_of_channels_groundtruth_table = []

    def load_config(self):

        # print(self.config_path)
        with open(self.config_path, 'r') as file:
            config = json.load(file)

        return config

    def connect_to_database(self):
        config = self.load_config()
        try:
            self.connection = mysql.connector.connect(
                host=config['host'],
                user=config['user'],
                password=config['password'],
                database=config['database'],
                port=config['port']
            )

            if self.connection.is_connected():
                print(f"Connected to {config['database']} database")
        except Error as e:
            print(f"Error connecting to {config['database']} database:", e)

    def disconnect(self):
        if self.connection.is_connected():
            self.connection.close()
            print("Closed connection to database")

    def execute_sql_file(self, file_path):
        try:

            config = self.load_config()

            # Establishing a connection to the database
            connectio = mysql.connector.connect(
                host=config['host'],
                database='allgovision',
                user=config['user'],
                password=config['password'],
                port=config['port']
            )

            # Reading the SQL file
            with open(file_path, 'r') as file:
                sql_commands = file.read()

            # Creating a cursor object and executing the SQL commands
            cursor = connectio.cursor()
            for command in sql_commands.split(';'):
                if command.strip():
                    cursor.execute(command)

            # Committing the changes to the database
            connectio.commit()

            print("SQL file executed successfully")

        except Error as e:
            print(f"Error: {e}")

    def list_of_channels_from_groundtruth_table(self):
        try:
            config = self.load_config()
            # Establish the database connection
            connection = mysql.connector.connect(host=config['host'],
                                                 database=config['database'],
                                                 user=config['user'],
                                                 password=config['password'],
                                                 port=config['port'])

            if connection.is_connected():
                cursor = connection.cursor()
                # SQL query to retrieve channels from the testcase table
                query = "SELECT test_case_name FROM groundtruth"
                cursor.execute(query)

                # Fetch all rows from the query
                results = cursor.fetchall()

                # Extract channels from the results and store them in self.testcase_channel_list
                self.list_of_channels_groundtruth_table = [
                    row[0] for row in results]
                # print(self.list_of_channels_groundtruth_table)
                connection.commit()
                # print(cursor.rowcount,
                #       "Record(s) inserted successfully into groundtruth table")

        except Error as e:
            print("Error while connecting to MariaDB", e)

    def get_list_of_mapped_channels_and_test_suit_name(self):
        try:
            cursor = self.connection.cursor()
            # SQL query to retrieve channels from the testcase table
            query = "SELECT test_case_name, test_suit_name FROM testcase"
            cursor.execute(query)

            # Fetch all rows from the query
            results = cursor.fetchall()

            self.testcase_channel_and_test_suit_list = [
                (row[0], row[1]) for row in results]
        except mysql.connector.Error as err:
            print(f"Error: {err}")
            self.testcase_channel_and_test_suit_list = []

    def insert_into_groundtruth(self, data):
        if (data[0] not in self.list_of_channels_groundtruth_table):
            try:
                # # Establish the database connection
                # connection = mysql.connector.connect(host=self.db_config['host'],
                #                                      database=self.db_config['database'],
                #                                      user=self.db_config['user'],
                #                                      password=self.db_config['password'],
                #                                      port=self.db_config['port'])

                if self.connection.is_connected():
                    cursor = self.connection.cursor()
                    insert_query = """
                        INSERT INTO groundtruth (
                            test_case_name, alarm_id, time_of_alarm, name_of_alarm,
                            alarm_metadata, object_metadata, alarm_image, object_crop
                        ) VALUES (%s, NULL, %s, %s, NULL, NULL, NULL, NULL)
                    """

                    cursor.execute(insert_query, data)
                    self.connection.commit()
                    # print(cursor.rowcount,
                    #       "Record(s) inserted successfully into groundtruth table")

            except Error as e:
                print("Error while connecting to MariaDB", e)

    def check_and_read_sheets_groundtruth(self, file_path, sheet_name):
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Get all sheet names
        sheet_names = workbook.sheetnames
        # print(sheet_names)

        sheet = workbook[sheet_name]

        channel_name = sheet_name
        # print(f"\nReading from sheet: {sheet_name}")
        # channel_name = sheet.cell(row=1, column=2).value

        # Example: Read values from the first column
        for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, min_col=2, max_col=3):
            combined_tuple = (channel_name, row[0].value, row[1].value)
            self.insert_into_groundtruth(combined_tuple)
        self.list_of_channels_groundtruth_table .append(channel_name)

        # Close the workbook to release memory
        workbook.close()
        print("Data inserted successfully into groundtruth table")

    def insert_into_testcase(self, data):
        name = (data[0], data[6])

        # print(self.testcase_channel_and_test_suit_list)
        if (name not in self.testcase_channel_and_test_suit_list):
            try:

                cursor = self.connection.cursor()
                insert_query = """
                    INSERT INTO testcase (
                        test_case_name, video_path, video_length_in_seconds, 
                        Description, configInfoPath, feature_name,test_suit_name
                    ) VALUES ( %s, %s, %s, %s, %s, %s,%s)
                """
                cursor.execute(insert_query, data)
                self.connection.commit()

                # print(cursor.rowcount,
                #       "Record(s) inserted successfully into testcase table")
            except Error as e:
                print("Error inserting into testcase:", e)

    def read_testcase_from_excel(self, file_path, sheet_name):
        # Load the workbook and select the sheet
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Iterate through the rows starting from the second row
        for row in sheet.iter_rows(min_row=2, values_only=True, min_col=2, max_col=8):
            # Skip rows where all values are None
            if all(cell is None for cell in row):
                continue

            # print(row)
            # Insert the records into the testcase table
            self.insert_into_testcase(row)
            channel_name = row[0]
            # Check if a sheet with this channel_name exists
            if channel_name in workbook.sheetnames:
                print(f"Sheet with name '{channel_name}' exists.")
                # You can now work with this sheet if needed
                # channel_sheet = workbook[channel_name]
                # Process the channel_sheet as required
                self.check_and_read_sheets_groundtruth(
                    file_path, channel_name)
            else:
                print(
                    f"Groundtruth table for the chanmmel  '{channel_name}' does not exist.")

        print("Record inserted successfully into test_case table")

   


def main():
    # Define the config file path and Excel file details
    config_path = 'config.json'
    file_path = 'Automation_details.xlsx'

    # Create an instance of AlarmSystemDatabase
    alarm_system_db = AlarmSystemDatabase(config_path)

    alarm_system_db.execute_sql_file('automation_testing.sql')

    # Connect to the database
    alarm_system_db.connect_to_database()
    alarm_system_db.list_of_channels_from_groundtruth_table()
    # alarm_system_db.list_of_channels_from_testcase_table()
    alarm_system_db.get_list_of_mapped_channels_and_test_suit_name()
    # alarm_system_db.list_of_channels_from_testrun_table()

    # Read data from Excel and insert into the database
    alarm_system_db.read_testcase_from_excel(file_path, 'Testcases')
    # alarm_system_db.read_testrun_from_excel(file_path, 'Testrun')

    # Disconnect from the database
    alarm_system_db.disconnect()


if __name__ == "__main__":
    main()
